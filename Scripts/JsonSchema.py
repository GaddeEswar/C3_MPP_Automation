import json
import sqlite3
import re
import os
import pdfplumber
import pandas as pd
from jsonschema import Draft7Validator, FormatChecker
from MainModule import JsonOperations,APIOperations
from SQLite import SQLiteConnection
from datetime import datetime,date
from openpyxl.styles import Alignment, PatternFill
from openpyxl import load_workbook
from pathlib import Path


class C3_MPP_JsonSchema():
    def __init__(self,BackupJson,ProjectJson,Product,Mode,AutomationSchemaPath):
        self.Header = {}
        self.Header['Product'] = Product
        self.Header['Mode'] = Mode
        self.BKjsonPath=BackupJson
        BKjson = JsonOperations(BackupJson)
        self.BKjsonData = BKjson.read_file()
        AutomationJsonSchema = JsonOperations(AutomationSchemaPath)
        SoftwareJsonSchema = JsonOperations(ProjectJson)
        self.AutomationJsonSchemaData = AutomationJsonSchema.read_file()
        self.SoftwareJsonSchemaData = SoftwareJsonSchema.read_file()
         # Connect to the SQLite database
        self.Conn =  sqlite3.connect('Resources/TestDataFile.db', check_same_thread=False)
        self.Japi = JsonOperations('json/Xpath.json')
        JapiDatatemp =self.Japi.read_file()
        self.JapiData = JapiDatatemp['API']
       
        
    def validate_with_exceptions(self):
        self.Logs=[]
        validator = Draft7Validator(self.AutomationJsonSchemaData, format_checker=FormatChecker())
        errors = sorted(
            validator.iter_errors(self.SoftwareJsonSchemaData),
            key=lambda e: list(e.path)
        )
      

        self.Logs.append('')
        self.Logs.append(str('- -' *20+ ' Schema Compliance Report ' + '- -' * 20))
        self.Logs.append('')

        # Build map of failed paths to their error messages
        failed_paths = {}
        if errors:
            for error in errors:
                if error.message == "None is not of type 'object'":
                    continue
                path = tuple(error.path)
                # For missing required property errors, map to the specific missing property
                if error.validator == "required":
                    match = re.search(r"'([^']+)' is a required property", error.message)
                    if match:
                        failed_paths[path + (match.group(1),)] = error.message
                    else:
                        failed_paths[path] = error.message
                elif error.validator == "additionalProperties":
                    match = re.search(r"'([^']+)' was unexpected", error.message)
                    if match:
                        failed_paths[path + (match.group(1),)] = error.message
                    else:
                        failed_paths[path] = error.message
                else:
                    failed_paths[path] = error.message

        # print("Failed Paths:", failed_paths)

        def path_has_error(prefix_tuple):
            for path in failed_paths:
                if len(path) >= len(prefix_tuple) and path[:len(prefix_tuple)] == prefix_tuple:
                    return failed_paths[path]
            return None

        def log_structure(current_path, current_val):
            self.Logs.append('')
            if isinstance(current_val, dict):
                for k, v in current_val.items():
                    sub_path = current_path + (k,)
                    err = path_has_error(sub_path)
                    indent = "  " * len(sub_path)
                    if err and "Additional properties are not allowed" in err and k in err:
                        self.Logs.append(f" {indent} SECTION:  {k} -- Additional properties are not allowed --- Fail")
                        continue
                  
                    if isinstance(v, dict):
                       
                        if err:
                            self.Logs.append(f" {indent} SECTION: {k} -- Did not followed the Schema ")
                        else:
                            self.Logs.append(f"{indent} SECTION: {k}")
                        log_structure(sub_path, v)
                    else:
                      
                        if err:
                            self.Logs.append(f"{indent}         {k}: {err} -- Fail")
                        else:
                            if k=='TestScope':self.Logs.append(f"{indent}         {k} :  {self.get_schema_type_for_path(sub_path)} -- Pass")

                            else:self.Logs.append(f"{indent}         {k} : {self.get_software_value_for_path(sub_path)}  -> {self.get_schema_type_for_path(sub_path)} -- Pass")
                            
                # Check for missing required properties at this level
                for path, err_msg in failed_paths.items():
                    if len(path) == len(current_path) + 1 and path[:len(current_path)] == current_path:
                        missing_key = path[-1]
                        if missing_key not in current_val:
                            indent = "  " * len(path)
                            self.Logs.append(f"{indent}      {missing_key} (Missing): {err_msg} --- Fail")

        # Create a dict having common Issue of TestScope
        TestScope_error = {}

        # Check for missing top level keys 
        for path, err_msg in failed_paths.items():
            if path and path[0] == 'TestingScope':
                TestScope_error.setdefault(err_msg, []).append(path)
                
            if len(path) == 1:
                missing_key = path[0]
                # if missing_key == 'TestingScope':
                #     continue
                if missing_key not in self.SoftwareJsonSchemaData:
                    self.Logs.append(f" {missing_key} (Missing): {err_msg} --- Fail")

        # Traverse and log status of all present keys and subfields (skipping TestingScope)
        for key in self.SoftwareJsonSchemaData.keys():
            if key != 'TestingScope':
                
                val = self.SoftwareJsonSchemaData[key]
                err = path_has_error((key,))
                if err and "Additional properties are not allowed" in err and key in err:
                    self.Logs.append(f"SECTION:  {key} -- Additional properties are not allowed --- Fail")
                    continue
                if isinstance(val, dict):
                    if err:
                        self.Logs.append(f"SECTION:  {key} -- Did not followed the Schema")
                    else:
                        self.Logs.append(f"SECTION:  {key}")
                    log_structure((key,), val)
                else:
                    if err:
                        self.Logs.append(f" [FAIL] --  {key} : {err} ")
                    else:
                        self.Logs.append(f" [PASS]  --  {key} : {self.get_software_value_for_path((key,))}  -> {self.get_schema_type_for_path((key,))} ")
                self.Logs.append('')
            else:
                self.Logs.append('- -' * 20 + ' SECTION : Testing Scope Verification ' + '- -' * 20)
                testing_scope = self.SoftwareJsonSchemaData.get('TestingScope', [])
                for msg, paths in TestScope_error.items():
                    self.Logs.append('')
                    Tests = {}
                    for path in paths:
                        idx = path[1]
                        if idx < len(testing_scope):
                            tc = testing_scope[idx]
                            test_id = str(tc['TestId'] + tc['TestName'])
                            Tests[test_id] = None

                    PathDisplay = ''.join(f'->{r}' for r in paths[0] if isinstance(r, str)) if paths else ''
                    self.Logs.append(f' Schema Path -- ↓↓')
                    self.Logs.append('')
                    self.Logs.append(f'            {PathDisplay}')
                    self.Logs.append('')
                    self.Logs.append(' Error Message -- ↓↓ ')
                    self.Logs.append('')
                    self.Logs.append(f'            {msg}')
                    self.Logs.append('')
                    self.Logs.append('-             Test Cases Affected  -- ↓↓            ')
                    self.Logs.append('')
                    for count, Test in enumerate(Tests, 1):
                        self.Logs.append(f'            {count} . {Test}')
                        self.Logs.append('')
                    
                    self.Logs.append('-' * 30 + ' ' + '-' * 30)
        
        # Write logs to a text file
        try:
            now = datetime.now()
            timestamp = now.strftime("%d%m%Y_%H%M%S")
            output_file=  f'Results/C3_MPP Excel Results/JsonSchem_Comparison_{self.Header['Product'] }_{self.Header['Mode'] }_{timestamp}.txt'
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(self.Logs))
        except Exception as e:
            print(f"Error writing SchemaReport.txt: {e}")

        return self.Logs

        
    def get_schema_type_for_path(self, path):
        """
        Traverses the automation JSON schema along the given path tuple/list
        and returns the defined 'type' of the field, if present.
        """
        curr = self.AutomationJsonSchemaData
        for segment in path:
            if isinstance(segment, str):
                if isinstance(curr, dict) and 'properties' in curr and segment in curr['properties']:
                    curr = curr['properties'][segment]
                else:
                    return None
            elif isinstance(segment, int):
                if isinstance(curr, dict) and 'items' in curr:
                    curr = curr['items']
                    if isinstance(curr, list):
                        if segment < len(curr):
                            curr = curr[segment]
                        else:
                            return None
                else:
                    return None
            else:
                return None
        
        if isinstance(curr, dict):
            return curr.get('type')
        return None
        
    def get_software_value_for_path(self, path):
        """
        Traverses the software JSON data along the given path tuple/list
        and returns the value, if present.
        """
        curr = self.SoftwareJsonSchemaData
        for segment in path:
            if isinstance(curr, dict) and segment in curr:
                curr = curr[segment]
            elif isinstance(curr, list) and isinstance(segment, int):
                if 0 <= segment < len(curr):
                    curr = curr[segment]
                else:
                    return None
            else:
                return None
        return curr
       
    def SegregateProjectFolder(self):
        self.TClist={}
       
        for TCdata in  self.BKjsonData['testBkpTestResultsandPath']:
            tracepathlist = TCdata['actualTracePath'].split('\\')
            ExactPath= self.BKjsonPath.replace('\\', '/').rsplit('/', 1)[0]# Remove the last segment
            tracepath = os.path.join(ExactPath,tracepathlist[len(tracepathlist)-3],tracepathlist[len(tracepathlist)-2],tracepathlist[len(tracepathlist)-1])
            TestId,TestName,TestResult=TCdata['testinformation']['TestId'],TCdata['testinformation']['TestName'],TCdata['testinformation']['TestResult']
            CoilType=TCdata['testinformation']['TesterConfiguration']['CoilUsed'].replace("#","_")
            # Get Coil details based on the coilType
            CoilQuery= f'''
                        SELECT 
                            CoilType.COIL_VOLTAGE,
                            CoilType.COIL_MODULATION,
                            CoilType.COIL_LOAD,
                            CoilType.PT_PHASE_LOAD
                        FROM CoilType WHERE CoilType.COIL_TYPE=='{CoilType}'
                        '''
            res=self.FetchDataFromQRY(CoilQuery)
            if res is not None:
                for _, row in res.iterrows():
                    insert_query = f"""
                        INSERT OR REPLACE INTO JsonDetails (JTestID, JTestName, JCoilType,JCoilVoltage,JCoilModulation,JCoilLoad,JPtPhaseLoad,TestResult,JTracepath)
                        VALUES ('{TestId}', '{TestName}', '{CoilType}','{row['COIL_VOLTAGE']}','{row['COIL_MODULATION']}','{row['COIL_LOAD']}','{row['PT_PHASE_LOAD']}','{TestResult}','{str(tracepath)}')
                    """
                    self.ExecutebyQuery(insert_query)
            
        # self.ExtractFromTrace()


    def ExtractFromTrace(self):

        Results={}
        TestCount=0
        for Test in self.SoftwareJsonSchemaData['TestExecutionDetails']['TestScope']: 
            if Test not in Results:Results[Test]={}

            GetTestDetailsQuery=f'''   SELECT * FROM JsonDetails WHERE JTestName='{Test}' '''
            res=self.FetchDataFromQRY(GetTestDetailsQuery)
            
            if res is not None:

                self.TraceUPL = APIOperations(url=self.JapiData[self.Header['Product']][self.Header['Mode']]['PutWaveformFile'])    
                self.TraceUPL.files = {"WaveformFile":open(res['JTracePath'][0].replace('/','\\'),"rb")}
                status = self.TraceUPL.PutRequest()
                if status == 200:
                    # print(f'{Test} trace is loaded')
                     #Get Packets___________________________________________________________________________
                    self.PktAPI = APIOperations(url=self.JapiData[self.Header['Product']][self.Header['Mode']]['GetCCLinePackets'],retype='json')
                    self.file_list = self.PktAPI.GetRequest()
                    id=0
                    PktCount=0
                    if 'TestingScope' not in Results[Test]:Results[Test]['TestingScope']={}
                    if 'TestLogs' not in Results[Test]['TestingScope']:Results[Test]['TestingScope']['TestLogs']={}
                    if 'TesterConfiguration' not in Results[Test]['TestingScope']:Results[Test]['TestingScope']['TesterConfiguration']={}

                    
                    try:
                        while id < len(self.file_list):
                            Type=self.GetPacketType(id)
                            if Type in ['Packet','Response']:
                                raw_data = [int(x, 16) for x in self.file_list[id]['header_Payload']['sFieldType'].split(':')[1].split()]
                                Pktduration=round((self.file_list[id]['stopTime']-self.file_list[id]['startTime'])*1000,1)
                                PktDescription = self.PktDescription(id)  
                                # print(PktDescription) 
                                if id not in Results[Test]['TestingScope']['TestLogs']:Results[Test]['TestingScope']['TestLogs'][id]={}
                                if self.SoftwareJsonSchemaData['TestingScope'][TestCount]['TestLogs'][PktCount]['PacketDescription']!=PktDescription:
                                    # print(PktDescription) 
                                    Results[Test]['TestingScope']['TestLogs'][id]['PacketDescription']=f'Received PktDesc from Trace is {PktDescription} and Json Description is { self.SoftwareJsonSchemaData['TestingScope'][TestCount]['TestLogs'][PktCount]['PacketDescription']}'
                                if round(self.SoftwareJsonSchemaData['TestingScope'][TestCount]['TestLogs'][PktCount]['PacketDuration'],1) !=Pktduration:
                                    Results[Test]['TestingScope']['TestLogs'][id]['PacketDuration']=f'Received PktDuration from Trace is {Pktduration} and Json Duration is {round(self.SoftwareJsonSchemaData['TestingScope'][TestCount]['TestLogs'][PktCount]['PacketDuration'],1)}'
                                if self.SoftwareJsonSchemaData['TestingScope'][TestCount]['TestLogs'][PktCount]['RawData']!=raw_data:
                                    Results[Test]['TestingScope']['TestLogs'][id]['RawData']=f'Received RawData from Trace is {raw_data} and Json RawData is { self.SoftwareJsonSchemaData['TestingScope'][TestCount]['TestLogs'][PktCount]['RawData']}'
                                PktCount+=1
                                if Results[Test]['TestingScope']['TestLogs'][id]=={}:del Results[Test]['TestingScope']['TestLogs'][id]
                            id+=1
                    except Exception as e: print(e)
            else:Results[Test]="Test did not found in DataBase"
            TestCount+=1
            
        
        with open('JsonResults.json', 'w') as json_file:
            json.dump(Results, json_file, indent=4)

    def PktDescription(self,id):

        pkt= self.file_list[id]['pktType']

        if 'SADT' in pkt:
            match = re.match(r'SADT/(\d+)([eo])', pkt)
            if match:
                number = match.group(1)
                parity = "even" if match.group(2) == 'e' else "odd"
                return f"Simultaneous_Auxiliary_Data_Transport_{number}_{parity} "
            
        elif 'PLA_2' in pkt:
            print(self.file_list[id]['pktType'])
            return f'Power_Loss_Accounting_2 {self.file_list[id]['value']}'
        elif "MATEDQ" in pkt:
            print(self.file_list[id]['pktType'])
            if  "MATEDQ-COEFF [0xA8]" in pkt:return f'Mated_Q_Coefficients {self.file_list[id]['value']}'
            elif "MATEDQ_RES [0x40]"in pkt:return f'Mated_Q_Result {self.file_list[id]['value']}'
        
        elif pkt in ["MODECAP [0x5A]"]: return f"Power_Mode_Capabilities - {self.file_list[id]['value']}"

        elif pkt in ["MODEXCAP [0xA0]"]: return "Power_Mode_Extended_Capabilities "

        elif pkt in ['MSR']:return f'Mode_Select_Request - {self.file_list[id]['value']}'
        elif pkt in ['MSS']: return f'Mode_Select_Status - {self.file_list[id]['value']}'

        elif pkt not in ['SADC','SDSR','DSR',"SRQ [0x20] "]:
            if pkt in ['Signal strength']: return f'{pkt.replace(' ','_').replace(':','_').title()} {self.file_list[id]['value']}'
            return f'{pkt.replace(' ','_').replace(':','_')} {self.file_list[id]['value']}'
       
        else:
            if pkt=='SADC':
                pkt='Simultaneous_Auxilory_Data_Control'
            elif pkt=='SDSR':
                pkt='Simultaneous_Data_Stream_Response'
            elif pkt=='DSR':
                pkt='Data_Stream_Response'
            elif pkt=="SRQ [0x20] ":
                pkt='Specific_Request'
        return f'{pkt} {self.file_list[id]['value']}'                        


    #1.Get packet Type, testermsg/packet/response
    def GetPacketType(self,id):
        if self.Header['Product'] == "C3":
            if self.Header['Mode'] == 'TPT':
                if self.file_list[id]['isTesterPkt']==False and self.file_list[id]['isFWTestermessage']==False:
                    return 'Packet'
                elif self.file_list[id]['isTesterPkt']==True and self.file_list[id]['isFWTestermessage']==False:
                    return 'Response'
                elif self.file_list[id]['isTesterPkt']==True and self.file_list[id]['isFWTestermessage']==True:
                    return 'TesterMsg'
            elif self.Header['Mode'] =="TPR":
                if self.file_list[id]['isTesterPkt']==True and self.file_list[id]['isFWTestermessage']==False:
                    return 'Packet'
                elif self.file_list[id]['isTesterPkt']==False and self.file_list[id]['isFWTestermessage']==False:
                    return 'Response'
                elif self.file_list[id]['isTesterPkt']==True and self.file_list[id]['isFWTestermessage']==True:
                    return 'TesterMsg'
        elif self.Header['Product'] == "MPP":
            if self.Header['Mode']=="TPR":
                if self.file_list[id]['isTesterPkt']==True and self.file_list[id]['isFWTestermessage']==False:
                    return 'Packet'
                elif self.file_list[id]['isTesterPkt']==False and self.file_list[id]['isFWTestermessage']==False:
                    return 'Response'
                elif self.file_list[id]['isTesterPkt']==True and self.file_list[id]['isFWTestermessage']==True:
                    return 'TesterMsg'
            elif self.Header['Mode']=="TPT":
                if self.file_list[id]['isTesterPkt']==True and self.file_list[id]['isFWTestermessage']==False:
                    return 'Packet'
                elif self.file_list[id]['isTesterPkt']==False and self.file_list[id]['isFWTestermessage']==False:
                    return 'Response'
                elif self.file_list[id]['isTesterPkt']==True and self.file_list[id]['isFWTestermessage']==True:
                    return 'TesterMsg'
        return None
            
    def FetchDataFromQRY(self,QRY):
        try:
            QRY_DF = pd.read_sql_query(QRY, self.Conn)
            if QRY_DF.shape[0] > 0:
                # print(list(QRY_DF['Name']))
                return QRY_DF
            return None
        except Exception as e:
            print(e)
            return None
    def ExecutebyQuery(self, QRY, params=None):
        """
        Execute a SQL query with optional parameters.

        :param QRY:   SQL query string (may contain ? placeholders)
        :param params: tuple or list of parameters for placeholders
        """
        try:
            self.cursor = self.Conn.cursor()
            if params:
                self.cursor.execute(QRY, params)
            else:
                self.cursor.execute(QRY)
            self.Conn.commit()
        except sqlite3.Error as e:
            print(f"Error occurred: {e}")
        except Exception as e:
            print(e)

   
 
       
    
class C3_MPP_PdfSchema():
    def __init__(self,BackupJson,ProjectJson,PdfFile,Product,Mode):
        BKjson = JsonOperations(BackupJson)
        self.BKjsonData = BKjson.read_file()
        ProjJson = JsonOperations(ProjectJson)
        self.ProjJsonData=ProjJson.read_file()
        self.BKjsonData = BKjson.read_file()
        self.Reports=JsonOperations('json/ReportsCompare.json')
        self.ReportsData=self.Reports.read_file()
        self.PdfFile=PdfFile
        self.Product = Product
        self.Mode = Mode
         # Connect to the SQLite database
        self.SQLConn = SQLiteConnection()
        self.database_path = 'Resources/GRLDB.db'
        self.connection = sqlite3.connect(self.database_path)

    def FormatPDFReport(self):

        current_section = None
        current_block = None
        res={}

        with pdfplumber.open(self.PdfFile) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                lines = [l.strip() for l in text.splitlines() if l.strip()]
                count=0
                for line in lines:
                    count+=1
                    # Skip headers / footers
                    l = line.replace(" ", "").lower()
                    if l.startswith("grl") or l.startswith("thisreportis") or l.startswith("page:") or l.startswith("date") :
                        continue
                    # ---------------- SECTION HEADERS ----------------
                    if line in self.ReportsData['Inputs'].keys():
                        current_section = self.ReportsData['Inputs'][line]
                        if current_section not in res:res[current_section]={}
                        continue
                    # ---------------- TEST RESULTS ----------------
                    if current_section == "TestingScopeAndResults":
                        keyres=False
                        for key , value in self.ReportsData['TestResult'].items():
                            if line.startswith(key):
                                current_block=value
                                if current_block not in res[current_section]:res[current_section][current_block]=[]
                                keyres=True
                                break
                        if keyres:continue
                        if  not  line.endswith("NONE"):
                            res[current_section][current_block].append(line)
                        continue
                    # ---------------- KEY : VALUE ----------------
                    if ":" in line and current_section:
                        key, value = map(str.strip, line.split(":", 1))
                        res[current_section][key.replace(" ", "")] = value if value else None 
                    if 'PTxInitiated' in line or 'EndOfCharge' in line or 'ForeignObjectDetection' in line: 
                        if current_section =="DeviceUnderTest":
                            key, value = map(str.strip, lines[count-2].split(":", 1))
                            res[current_section][key.replace(" ", "")] = str(value)+str(line) if value else None 


                   

        # print(res)
        with open("out.json", "w", encoding="utf-8") as f:
            json.dump(res, f, indent=2, ensure_ascii=False)
        self.PdfSegregate(res)


    def PdfSegregate(self,PdfReportJson={}):

        self.PdfResult={}

        for key in self.ReportsData['C3_MPP'][self.Mode]['PDF']:
            
            if key=="TestingScopeAndResults":continue
            if key not in PdfReportJson:
                self.PdfResult[key]="MissingKey"
                continue
            else:self.PdfResult[key]={}
            for FieldName in self.ReportsData['C3_MPP'][self.Mode]['PDF'][key]:
                if FieldName not in PdfReportJson[key]:self.PdfResult[key][FieldName]="MissingField"
                else:self.PdfResult[key][FieldName]={
                    "PDF":PdfReportJson[key][FieldName],
                    "Json":'',
                    "Result":"Pass"
                }

        for key in PdfReportJson:
            try:
                if key=="TestingScopeAndResults":continue
                if key not in self.PdfResult:
                    self.PdfResult[key]="ExtraKey"
                    continue
                for FieldName in PdfReportJson[key]:
                    if FieldName not in self.PdfResult[key]:self.PdfResult[key][FieldName]="ExtraField"

                    if key in self.ReportsData['C3_MPP'][self.Mode]['JsonReferenceKeys'] and FieldName in self.ReportsData['C3_MPP'][self.Mode]['JsonReferenceKeys'][key] and self.PdfResult[key][FieldName]!="ExtraField":
                        if  self.PdfResult[key][FieldName]=="MissingField":continue  
                        elif key=="Overview":self.ValidateOverivew(key,FieldName)   
                        else: self.PdfValidate(key,FieldName)
            except Exception as e:print(e)
        
        # with open("Pdfres.json", "w", encoding="utf-8") as f:
        #     json.dump(self.PdfResult, f, indent=2, ensure_ascii=False)
        

    def PdfValidate(self,key,FieldName):

        #Validation
        try:
            data=self.getBckupdata(key,FieldName)
            if data is not None: 
                self.PdfResult[key][FieldName]['Json']=data
                PdfVal=self.PdfResult[key][FieldName]['PDF'] 
                JsonVal= self.PdfResult[key][FieldName]['Json'] 
                Result='Pass'
                # Validate based on data Type
                Types=self.ReportsData['C3_MPP'][self.Mode]['JsonReferenceKeys'][key][FieldName][2]
                if  Types[1]=="int":
                    PdfVal=PdfVal.replace('W','') if FieldName.lower().endswith('power') else PdfVal
                    if (PdfVal is None and JsonVal is not None )or (int(PdfVal) != JsonVal): Result='Fail'
                elif Types[1]=="Boolean":
                    if PdfVal in ['Disabled','No'] and data : Result='Fail'  
                elif Types[1]=="List":
                    JsonVal=list(str(x).replace(" ","") for x in JsonVal) if len(JsonVal)>0 else ['None']
                    PdfVal= list(PdfVal.replace('.','').split(',')) if PdfVal is not None else ['None']
                    if  (PdfVal[0] =="None" and len(JsonVal)>1) or (PdfVal[0] !="None" and JsonVal[0]=='None'): Result='Fail'
                    else:
                        if PdfVal !=JsonVal: Result='Fail'    
                    if len(JsonVal)>0 :self.PdfResult[key][FieldName]['Json']=','.join(JsonVal)
                else:
                    if PdfVal is not None and PdfVal!=JsonVal:Result='Fail'
                   
                # Update Final Result
                self.PdfResult[key][FieldName]['Result']=Result

        except Exception as e: print(e)
                       
    def ValidateOverivew(self,key,FieldName):

        if FieldName in ['STARTED','FINISHED']:

            min_time = None
            max_time = None

            for Test in self.ProjJsonData['TestingScope']:
                start_time = datetime.fromisoformat(Test['TestStartTime'])
                end_time = datetime.fromisoformat(Test['TestEndTime'])
                if min_time is None or start_time < min_time: min_time = start_time               
                if max_time is None or end_time > max_time: max_time = end_time
                
            Jsonval= min_time if FieldName=="STARTED" else max_time
            self.PdfResult[key][FieldName]['Json']=Jsonval.strftime("%Y-%m-%d %H:%M:%S.%f")
            PdfVal = self.PdfResult[key][FieldName]['PDF'].replace('"', '').strip()
            if len(PdfVal) > 10 and PdfVal[10] != " ": PdfVal = PdfVal[:10] + " " + PdfVal[10:] 
            difference=abs((datetime.strptime(PdfVal, "%Y-%m-%d %H:%M:%S.%f") - ( Jsonval.replace(tzinfo=None) if Jsonval.tzinfo else Jsonval)).total_seconds())
            if difference > 70: self.PdfResult[key][FieldName]['Result']='Fail'

        else:
            TestResult=list(self.ProjJsonData['TestExecutionDetails']['TestResult'].split(','))
            Index_Type=self.ReportsData['C3_MPP'][self.Mode]['JsonReferenceKeys'][key][FieldName]
            Jsonval=int(TestResult[Index_Type[0]].split(':')[1])
            Total=int(TestResult[1].split(':')[1])
            # Total = sum(int(x.split(':')[1]) for x in TestResult[2:])
            if not Index_Type[1]:
                if int(self.PdfResult[key][FieldName]['PDF']) !=Jsonval:self.PdfResult[key][FieldName]['Result']='Fail'
                self.PdfResult[key][FieldName]['Json']=int(TestResult[1].split(':')[1])
            else:
                Percentage=round((Jsonval/Total)*100,2)
                if Percentage!=float(self.PdfResult[key][FieldName]['PDF'].replace('%','')):self.PdfResult[key][FieldName]['Result']='Fail'
                self.PdfResult[key][FieldName]['Json']=str(f'{Percentage}%')

    def PdfDB(self):
        self.SQLConn.ExecutebyQuery(f"DELETE FROM PdfReports")
        query = """
                    INSERT INTO PdfReports
                    (SeqNo,StandardReportkey, FieldSeqNo,StandardReportFieldName, SWReportKey, SWFieldName, SWFieldContent,JsonFieldContent,Result)
                    VALUES (?, ?, ?, ? , ?, ?, ?, ?, ?)
                """
        SeqNo=1
        for key in self.PdfResult:
            if type(self.PdfResult[key])==dict:
                FieldSeqNo=1
                for FieldName in self.PdfResult[key]:
                    Params=None
                    if type(self.PdfResult[key][FieldName])==dict:
                        PDFval= self.PdfResult[key][FieldName]['PDF']
                        Jsonval= self.PdfResult[key][FieldName]['Json']
                        params = (SeqNo, key, FieldSeqNo,FieldName,  key, FieldName,
                                 PDFval if PDFval is not None else "None", Jsonval if Jsonval is not None else "None",
                                 1 if self.PdfResult[key][FieldName]['Result']=='Pass' else 0 )
                    else:
                        if self.PdfResult[key][FieldName]=="ExtraField":  params = (SeqNo, key, FieldSeqNo,'NA',  key, FieldName,"","",1)  
                        else: params = (SeqNo, key, FieldSeqNo,FieldName,  key, 'NA',"","",0)      
                    self.SQLConn.ExecutebyQuery(query,params )
                    FieldSeqNo+=1
            else:
                params=None
                if self.PdfResult[key]=="ExtraKey": params = (SeqNo, 'NA', 0,"",  key, "","","",1)          
                else:  params = (SeqNo, key, 0,"",'NA', "","","",0) 
                self.SQLConn.ExecutebyQuery(query,params )
            SeqNo+=1
                                 
        #For Excel generation
        Header_Qry =""" SELECT
                        SeqNo as SL_No,FieldSeqNo as FieldNo,StandardReportKey,StandardReportFieldName,SWReportKey,SWFieldName,SWFieldContent,JsonFieldContent,Result 
                        FROM PdfReports
                        
                    """
        index=["SL_No",'StandardReportKey','SWReportKey','FieldNo','StandardReportFieldName','SWFieldName','SWFieldContent','JsonFieldContent','Result']
        return self.ExcelReport(Header_Qry,index,'PDF')

    def ExcelReport(self,Query,index=[],Type=''):

        Header_Qry =Query
        Header_df = pd.read_sql_query(Header_Qry, self.connection)
        Header_df['Result'] = Header_df['Result'].replace({1: "pass", 0: "fail"})

        # print(Header_df)
         # Create a pivot table
        pivot_table = pd.pivot_table(
            Header_df,
            index=index,  # Rows
            values=[],        # Values to aggregate
            aggfunc='sum'       # Aggregation function
        )
        now = datetime.now()
        timestamp = now.strftime("%d%m%Y_%H%M%S")
        output_file = f'Results/C3_MPP Excel Results/{Type}_Comparison_{self.Product}_{self.Mode}_{timestamp}.xlsx'
        with pd.ExcelWriter(output_file) as writer:
            pivot_table.to_excel(writer, sheet_name=f"{Type}Report")
        self.format_excel(output_file)
        return output_file

    def format_excel(self,file_path):
        # Define color fills for results (Global to avoid redundant creation)
        COLOR_MAPPING = {
            "pass": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Green
            "fail": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
            # "inconclusive": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Yellow
            "na": PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid"),
        }
        
        """Loads, processes, and saves the Excel file with formatting."""
        # Load the Excel file
        wb = load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]  # Get the active sheet
            # print("sheet_name:",sheet_name)

            # Auto-fit all columns based on content
            for col in ws.iter_cols():
                # print("col:",col)
                col_letter = col[0].column_letter  # Get column letter
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[col_letter].width = max_length + 2  # Add padding

            # Find the "Remarks" column dynamically
            remarks_col_index = next(
                (col[0].column for col in ws.iter_cols(1, ws.max_column) 
                if col[0].value and str(col[0].value).strip().lower() == "remarks"),
                None
            )

            # Align "Remarks" column to left if found
            if remarks_col_index:
                for cell in ws.iter_cols(min_col=remarks_col_index, max_col=remarks_col_index, 
                                        min_row=2, max_row=ws.max_row):
                    for c in cell:
                        c.alignment = Alignment(horizontal="left")

            # Apply color coding for Pass/Fail/Inconclusive
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value).strip().lower()
                        if cell_value in COLOR_MAPPING:
                            cell.fill = COLOR_MAPPING[cell_value]

        # Save the updated file
        wb.save(file_path)

    def getBckupdata(self,key,FieldName):
        try :
            BckupJson=self.ReportsData['C3_MPP'][self.Mode]['JsonReferenceKeys'][key][FieldName]
            if BckupJson[0]: data=self.BKjsonData                              
            else: data=self.ProjJsonData
            for vals in BckupJson[1]:
                data=data[vals]
            return data
        except Exception as e:
            print(e)
            return None



# obj=C3_MPP_JsonSchema(
#     AutomationSchemaPath=r"C:\Users\Eswar\Downloads\V230_GRL_C3_FinalReport (2).json",
#     SoftwareSchemaPath=r"C:\Users\Eswar\Downloads\V230_GRL_C3_FinalReport (2).json",
#     BackupJson=r"C:\Users\Eswar\Downloads\V221_Final_TestBackup (1).gproj",
#     Product='C3',Mode='TPR'
# )
# obj.ExtractTestFromDB()