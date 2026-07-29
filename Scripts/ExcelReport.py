import sqlite3
import pandas as pd
from datetime import datetime,date
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from SQLite import SQLiteConnection 

class ExcelReports():
    def __init__(self):
        #1. get the Data from SQL DB
        # Define the path to the SQLite database
        self.database_path = 'Resources/GRLDB.db'
        # Connect to the SQLite database
        self.connection = sqlite3.connect(self.database_path)
        self.SQLcon = SQLiteConnection()
    def CTSDetailedReport(self,filters,product):
        #Taks
        SW =  f"('{filters['SW'][0]}')" if len (filters['SW']) == 1 else tuple(filters['SW'])
        FW =  f"('{filters['FW'][0]}')" if len (filters['FW']) == 1 else tuple(filters['FW'])
        HW =  f"('{filters['HW'][0]}')" if len (filters['HW']) == 1 else tuple(filters['HW'])
        Board =  f"('{filters['Board'][0]}')" if len (filters['Board']) == 1 else tuple(filters['Board'])
        DUTname =  f"('{filters['DUTname'][0]}')" if len (filters['DUTname']) == 1 else tuple(filters['DUTname'])
        DUTID =  f"('{filters['DUTID'][0]}')" if len (filters['DUTID']) == 1 else tuple(filters['DUTID'])
        Chap =  f"('{filters['Chap'][0]}')" if len (filters['Chap']) == 1 else tuple(filters['Chap'])
        Coil =  f"('{filters['Coil'][0]}')" if len (filters['Coil']) == 1 else tuple(filters['Coil'])
        Tests =  f"('{filters['Tests'][0]}')" if len (filters['Tests']) == 1 else tuple(filters['Tests'])
        Checks =[]
        if filters['Timings'] == True:Checks.append('Timing')
        if filters['Measures'] == True:Checks.append('Measures')
        if filters['Timings'] == True:Checks.append('Others')
        checkslist = f"('{Checks[0]}')" if len(Checks)==1 else tuple(Checks)
        #1. Split excel sheet by phase wise
        # Phases = self.SQLcon.FetchDataFromQRY("SELECT DISTINCT(ChapterName) from Header")
        # Phaselist = Phases['ChapterName'] if Phases is not None else []
        FinalDataFrames = {}
        if len(filters['Chap'])>0:
            for pha in filters['Chap']:
                # print(pha)
                ph = f"('{pha}')"
                Header_Qry = f'''
                SELECT Header.TestcaseName, Header.SWresult,Header.TCresult as AutomationResult,ChecksHeader.Type as CheckType,
                ChecksHeader.SEQID as Flow,ChecksHeader.CheckSEQ,ChecksHeader.Description,ChecksDetails.CheckSEQ as SubID,
                ChecksDetails.Remarks,ChecksDetails.Result
                from Header 
                LEFT JOIN ChecksHeader on Header.UID = ChecksHeader.UID
                Left JOIN ChecksDetails on ChecksHeader.UID = ChecksDetails.UID and ChecksHeader.Description = ChecksDetails.Description and ChecksHeader.SEQID = ChecksDetails.SEQID
                WHERE
                Certification = "{filters['Certification']}" and BoardModel="{filters['Product']}" and  SWVersion in {SW} and FWVersion in {FW} and HWVersion in {HW} and BoardNo in {Board} and DUTName in {DUTname} and DUTID in {DUTID} and
                ChapterName in {ph} and Coil in {Coil} and TestcaseID in {Tests}
                ORDER by Header.TestcaseName,ChecksHeader.SEQID,ChecksHeader.SEQID, ChecksDetails.CheckSEQ
                '''
                # WHERE 
                # Certification = "{filters['Certification']}" and BoardModel="{filters['Product']}" and  SWVersion in {SW} and FWVersion in {FW} and HWVersion in {HW} and BoardNo in {Board} and DUTName in {DUTname} and DUTID in {DUTID} and
                # ChapterName in {Chap} and Coil in {Coil} and TestcaseID in {Tests}
                # ORDER by Header.TestcaseName,ChecksHeader.SEQID,ChecksHeader.SEQID, ChecksDetails.CheckSEQ
                # ChecksHeader.Type="Measures" and
                Header_df = pd.read_sql_query(Header_Qry, self.connection)
                if product=="MPP":
                    Header_df["Flow"] = Header_df["Flow"].replace("1", "128kHz")
                    Header_df["Flow"] = Header_df["Flow"].replace("2", "360kHz")
                else: Header_df["Flow"] = Header_df["Flow"]
                # print(Header_df)
                # Create a pivot table
                pivot_table = pd.pivot_table(
                    Header_df,
                    index=['TestcaseName','SWresult','AutomationResult','Flow','CheckType','CheckSEQ','Description','SubID','Remarks'],  # Rows
                    values=['Result'],        # Values to aggregate
                    aggfunc='sum'                        # Aggregation function
                )
                FinalDataFrames[pha]=pivot_table
                # print(pivot_table)
                # print(FinalDataFrames)
            
            now = datetime.now()
            timestamp = now.strftime("%d%m%Y_%H%M%S")
            output_file = f'Results/C3_MPP Excel Results/CTSChecks_Report_{filters['Product']}_{filters['Certification']}_{filters['SW'][0]}_{filters['FW'][0]}_{timestamp}.xlsx'
            with pd.ExcelWriter(output_file) as writer:
                for Phases in FinalDataFrames:
                    # print("Phases:",Phases)
                    phasename = Phases if len(Phases)<30 else Phases[0:30]
                    FinalDataFrames[Phases].to_excel(writer, sheet_name=phasename)
                # for sheet_name, df in FinalDataFrames.items():
                    # df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            #     pivot_table.to_excel(writer, sheet_name='Pivot Table')
            #Update Excel sheet
            self.format_excel(output_file)
    
    def CTSCompleteReport(self,product,Mode,filters):

        print(filters)
        if filters['SW'][0]=='':

            Header_Qry = f'''
            SELECT Header.TestcaseName, Header.SWresult,Header.TCresult as AutomationResult,ChecksHeader.Type as CheckType,
            ChecksHeader.SEQID as Flow,ChecksHeader.CheckSEQ,ChecksHeader.Description,ChecksDetails.CheckSEQ as SubID,
            ChecksDetails.Remarks,ChecksDetails.Result
            from Header 
            LEFT JOIN ChecksHeader on Header.UID = ChecksHeader.UID
            Left JOIN ChecksDetails on ChecksHeader.UID = ChecksDetails.UID and ChecksHeader.Description = ChecksDetails.Description and ChecksHeader.SEQID = ChecksDetails.SEQID
            ORDER by Header.TestcaseName,ChecksHeader.SEQID,ChecksHeader.SEQID, ChecksDetails.CheckSEQ
            '''
        
            Header_df = pd.read_sql_query(Header_Qry, self.connection)
            if product=="MPP":
                Header_df["Flow"] = Header_df["Flow"].replace("1", "128kHz")
                Header_df["Flow"] = Header_df["Flow"].replace("2", "360kHz")
            else: Header_df["Flow"] = Header_df["Flow"]
            # print(Header_df)
            # Create a pivot table
            pivot_table = pd.pivot_table(
                Header_df,
                index=['TestcaseName','SWresult','AutomationResult','Flow','CheckType','CheckSEQ','Description','SubID','Remarks'],  # Rows
                values=['Result'],        # Values to aggregate
                aggfunc='sum'                        # Aggregation function
            )
            now = datetime.now()
            timestamp = now.strftime("%d%m%Y_%H%M%S")
            output_file = f'Results/C3_MPP Excel Results/CTSChecks_Report_{product}_{Mode}_{timestamp}.xlsx'
            with pd.ExcelWriter(output_file) as writer:pivot_table.to_excel(writer, sheet_name="AllTestCases")
            #Update Excel sheet
            self.format_excel(output_file)

    def summarize_Report(self,filters,product):
        Total_test_case_count = 0
        test_case_details = []
        SW =  f"('{filters['SW'][0]}')" if len (filters['SW']) == 1 else tuple(filters['SW'])
        FW =  f"('{filters['FW'][0]}')" if len (filters['FW']) == 1 else tuple(filters['FW'])
        HW =  f"('{filters['HW'][0]}')" if len (filters['HW']) == 1 else tuple(filters['HW'])
        Board =  f"('{filters['Board'][0]}')" if len (filters['Board']) == 1 else tuple(filters['Board'])
        DUTname =  f"('{filters['DUTname'][0]}')" if len (filters['DUTname']) == 1 else tuple(filters['DUTname'])
        DUTID =  f"('{filters['DUTID'][0]}')" if len (filters['DUTID']) == 1 else tuple(filters['DUTID'])
        Chap =  f"('{filters['Chap'][0]}')" if len (filters['Chap']) == 1 else tuple(filters['Chap'])
        Coil =  f"('{filters['Coil'][0]}')" if len (filters['Coil']) == 1 else tuple(filters['Coil'])
        Tests =  f"('{filters['Tests'][0]}')" if len (filters['Tests']) == 1 else tuple(filters['Tests'])
 
        Checks =[]
        if filters['Timings'] == True:Checks.append('Timing')
        if filters['Measures'] == True:Checks.append('Measures')
        if filters['Timings'] == True:Checks.append('Others')
        checkslist = f"('{Checks[0]}')" if len(Checks)==1 else tuple(Checks)
        if len(filters['Chap'])>0:
            for pha in filters['Chap']:
                # print(pha)
                ph = f"('{pha}')"
                Header_Qry = f'''
                SELECT Header.TestcaseName, Header.SWresult,Header.TCresult as AutomationResult
                from Header
                WHERE
                Certification = "{filters['Certification']}" and BoardModel="{filters['Product']}" and  SWVersion in {SW} and FWVersion in {FW} and HWVersion in {HW} and BoardNo in {Board} and DUTName in {DUTname} and DUTID in {DUTID} and
                ChapterName in {ph} and Coil in {Coil} and TestcaseID in {Tests}
                ORDER by Header.TestcaseName
                '''
                Header_df = pd.read_sql_query(Header_Qry, self.connection)
                # pivot_table = Header_df  
                test_case_details.append(Header_df)
            # print(test_case_details)
            final_df = pd.concat(test_case_details, ignore_index=True)
            pivot_table = final_df
            print(final_df)
            total_test_Qry = f'''
                SELECT OfflineTestcases.TestCase
                from OfflineTestcases
                WHERE
                status = 0 OR status = 1
                ORDER by OfflineTestcases.TestCase
                '''          
            offline_test_cases_df = pd.read_sql_query(total_test_Qry, self.connection)
            # print(Header_df['TestcaseName'])
            # print(pivot_table)
            # print(offline_test_cases_df)
            Total_test_case_count = len(offline_test_cases_df)
            Selected_Test_case_count = 0
            Pass_count = 0
            Inconclusive_count = 0
            Fail_count = 0
            NA_count = 0
            result_count = {'Total_test_case_count': Total_test_case_count,'Selected_Test_case_count' : 0, 'Pass_count' : 0, 'Fail_count' : 0, 'Inconclusive_count' : 0, 'NA_count': 0}
            for test_case, auto_res in zip(pivot_table['TestcaseName'], pivot_table['AutomationResult']):
                Selected_Test_case_count += 1
                # result_count['test_cases'].append(test_case)
                if auto_res == "Pass":
                    Pass_count += 1
                    result_count['Pass_count'] = Pass_count
                elif auto_res == "Fail":
                    Fail_count += 1
                    result_count['Fail_count'] = Fail_count
                elif auto_res == "Inconclusive":
                    Inconclusive_count += 1
                    result_count['Inconclusive_count'] = Inconclusive_count
                else:
                    NA_count += 1
                    result_count['NA_count'] = NA_count
            result_count['Selected_Test_case_count'] = Selected_Test_case_count
            # print(result_count)          
            now = datetime.now()
            timestamp = now.strftime("%d%m%Y_%H%M%S")
            output_file = (f'Results/C3_MPP Excel Results/_{'Summary_table'}_{filters['Product']}_{filters['Certification']}_{filters['SW'][0]}_{filters['FW'][0]}_{timestamp}.xlsx')
            with pd.ExcelWriter(output_file, engine="xlsxwriter") as writer:
                workbook = writer.book
                worksheet = workbook.add_worksheet("Summary Table")
                writer.sheets["Summary Table"] = worksheet
                row = 0
                header_format = workbook.add_format({'bold': True,'font_size': 14})
                worksheet.write(row, 0 ,"Test_case_Summary", header_format)
                row += 2
                for key, value in result_count.items():
                    worksheet.write(row, 0, key)
                    if isinstance(value, list):
                        worksheet.write(row, 1, str(value))
                    else:
                        worksheet.write(row, 1, value)
                    row += 1
                pivot_table.insert(0, "S.No", range(1, len(pivot_table) + 1))
                pivot_table.to_excel(
                    writer,
                    sheet_name="Summary Table",
                    startrow=row,
                    index=False
                )
                worksheet_obj = writer.sheets["Summary Table"]    
                pass_format = workbook.add_format({'bg_color': "#13EC3E"})        
                fail_format = workbook.add_format({'bg_color': "#F0102A"})        
                inconclusive_format = workbook.add_format({'bg_color': "#EA7F0E"})
                na_format = workbook.add_format({'bg_color': "#4371D5"})          
                columns_to_color = ["AutomationResult", "SWresult"]
                for col_name in columns_to_color:
                    if col_name in pivot_table.columns:
                        col_idx = list(pivot_table.columns).index(col_name)
                        start_data_row = row + 1
                        for i, value in enumerate(pivot_table[col_name]):
                            excel_row = start_data_row + i
                            if value == "Pass":
                                fmt = pass_format
                            elif value == "Fail":
                                fmt = fail_format
                            elif value == "Inconclusive":
                                fmt = inconclusive_format
                            else:
                                fmt = na_format
                            worksheet_obj.write(excel_row, col_idx, value, fmt)
               
                for col_num, col_name in enumerate(pivot_table.columns):
                    max_len = len(col_name)
                    max_len = max(max_len, pivot_table[col_name].astype(str).map(len).max())
                    if col_num < 2:  
                        summary_values = [str(k) for k in result_count.keys()] + \
                                        [str(v) for v in result_count.values()]
                        max_len = max(max_len, max(len(x) for x in summary_values))
                    worksheet_obj.set_column(col_num, col_num, max_len + 2)
    
    def CTSChecksReport(self,filters):
        SW =  f"('{filters['SW'][0]}')" if len (filters['SW']) == 1 else tuple(filters['SW'])
        FW =  f"('{filters['FW'][0]}')" if len (filters['FW']) == 1 else tuple(filters['FW'])
        HW =  f"('{filters['HW'][0]}')" if len (filters['HW']) == 1 else tuple(filters['HW'])
        Board =  f"('{filters['Board'][0]}')" if len (filters['Board']) == 1 else tuple(filters['Board'])
        DUTname =  f"('{filters['DUTname'][0]}')" if len (filters['DUTname']) == 1 else tuple(filters['DUTname'])
        DUTID =  f"('{filters['DUTID'][0]}')" if len (filters['DUTID']) == 1 else tuple(filters['DUTID'])
        Chap =  f"('{filters['Chap'][0]}')" if len (filters['Chap']) == 1 else tuple(filters['Chap'])
        Coil =  f"('{filters['Coil'][0]}')" if len (filters['Coil']) == 1 else tuple(filters['Coil'])
        Tests =  f"('{filters['Tests'][0]}')" if len (filters['Tests']) == 1 else tuple(filters['Tests'])

        Checks =[]
        if filters['Timings'] == True:Checks.append('Timing')
        if filters['Measures'] == True:Checks.append('Measures')
        if filters['Timings'] == True:Checks.append('Others')
        checkslist = f"('{Checks[0]}')" if len(Checks)==1 else tuple(Checks)

        # Define your SQL query
        Header_Qry = f'''
        SELECT UID,SWVersion,FWVersion,HWVersion,BoardNo,BoardModel,Certification,DUTName,DUTID,Coil,TestcaseName,ChapterName,SWresult,TCresult FROM Header 
        WHERE Certification = "{filters['Certification']}" and BoardModel="{filters['Product']}" and  SWVersion in {SW} and FWVersion in {FW} and HWVersion in {HW} and BoardNo in {Board} and DUTName in {DUTname} and DUTID in {DUTID} and
        ChapterName in {Chap} and Coil in {Coil} and TestcaseID in {Tests}
        ORDER by UID
        '''
        Timing_Qry = f'''
        SELECT UID,SEQID,Description,Type,ExpValue,MinValue,MaxValue,Value,Result,Remarks from ChecksHeader 
        WHERE Type in {checkslist}
        ORDER by UID,Type,SEQID,CheckSEQ
        '''
        # Measures_Qry = '''
        # SELECT UID,SEQID,Description,ExpValue,MinValue,MaxValue,Result from ChecksHeader WHERE Type="Measures" ORDER by UID,SEQID
        # '''
        # Load the query result into a DataFrame
        Header_df = pd.read_sql_query(Header_Qry, self.connection)
        Timing_df = pd.read_sql_query(Timing_Qry, self.connection)
        # Measures_df = pd.read_sql_query(Measures_Qry, self.connection)
        # print(Header_df)

        # Close the connection
        # self.connection.close()
        #Merge checks
        merged_df = pd.merge(Header_df, Timing_df, on='UID')
        #Update remarks
        merged_df['Remarks']=merged_df['Remarks'].replace('','NA').fillna('NA')
        # print(merged_df)
        # Create a pivot table
        pivot_table = pd.pivot_table(
            merged_df,
            index=['Certification','SWVersion','FWVersion','HWVersion','BoardModel','BoardNo','ChapterName','TestcaseName','SWresult','SEQID', 'Type','Description','ExpValue'
                ,'Value','MinValue','MaxValue'],  # Rows
            values=['Remarks','Result'],        # Values to aggregate
            aggfunc='sum'                        # Aggregation function
        )
        # print(pivot_table)
        now = datetime.now()
        timestamp = now.strftime("%d%m%Y_%H%M%S")
        output_file = f'Results/C3_MPP Excel Results/CTSChecks_Report_{filters['Product']}_{filters['Certification']}_{filters['SW'][0]}_{filters['FW'][0]}_{timestamp}.xlsx'
        # Write both the original data and the pivot table to an Excel file
        with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            # Write the original data
            # pivot_table.to_excel(writer, sheet_name='Data', index=False)
            # Write the pivot table
            pivot_table.to_excel(writer, sheet_name='Pivot Table') 
    
    def PacketPayLoadDetailedReport(self,filters,product):
        #Taks
        SW =  f"('{filters['SW'][0]}')" if len (filters['SW']) == 1 else tuple(filters['SW'])
        FW =  f"('{filters['FW'][0]}')" if len (filters['FW']) == 1 else tuple(filters['FW'])
        HW =  f"('{filters['HW'][0]}')" if len (filters['HW']) == 1 else tuple(filters['HW'])
        Board =  f"('{filters['Board'][0]}')" if len (filters['Board']) == 1 else tuple(filters['Board'])
        DUTname =  f"('{filters['DUTname'][0]}')" if len (filters['DUTname']) == 1 else tuple(filters['DUTname'])
        DUTID =  f"('{filters['DUTID'][0]}')" if len (filters['DUTID']) == 1 else tuple(filters['DUTID'])
        Chap =  f"('{filters['Chap'][0]}')" if len (filters['Chap']) == 1 else tuple(filters['Chap'])
        Coil =  f"('{filters['Coil'][0]}')" if len (filters['Coil']) == 1 else tuple(filters['Coil'])
        Tests =  f"('{filters['Tests'][0]}')" if len (filters['Tests']) == 1 else tuple(filters['Tests'])
        # Checks =[]
        # if filters['Timings'] == True:Checks.append('Timing')
        # if filters['Measures'] == True:Checks.append('Measures')
        # if filters['Timings'] == True:Checks.append('Others')
        # checkslist = f"('{Checks[0]}')" if len(Checks)==1 else tuple(Checks)
        #1. Split excel sheet by phase wise
        # Phases = self.SQLcon.FetchDataFromQRY("SELECT DISTINCT(ChapterName) from Header")
        # Phaselist = Phases['ChapterName'] if Phases is not None else []
        FinalDataFrames = {}
        if len(filters['Chap'])>0:
            for pha in filters['Chap']:
                # print(pha)
                ph = f"('{pha}')"
                Header_Qry = f'''
                SELECT Header.TestcaseName,PayLoadDetails.SEQID as Flow,PayLoadDetails.Type as CheckType,PayLoadDetails.PacketID,
                PayLoadDetails.Packet,PayLoadDetails.HeaderName,PayLoadDetails.Byte,PayLoadDetails.Bit,PayLoadDetails.CheckName ,
                PayLoadDetails.ExpValue,PayLoadDetails.RecValue ,PayLoadDetails.ChecksResult,PayLoadDetails.HeaderResult
                from Header 
                LEFT JOIN PayLoadDetails on Header.UID = PayLoadDetails.UID
                WHERE
                Certification = "{filters['Certification']}" and BoardModel="{filters['Product']}" and  SWVersion in {SW} and FWVersion in {FW} and HWVersion in {HW} and BoardNo in {Board} and DUTName in {DUTname} and DUTID in {DUTID} and
                ChapterName in {ph} and Coil in {Coil} and TestcaseID in {Tests}
                ORDER by Header.TestcaseName,PayLoadDetails.SEQID,PayLoadDetails.PacketID
                '''
                # WHERE 
                # Certification = "{filters['Certification']}" and BoardModel="{filters['Product']}" and  SWVersion in {SW} and FWVersion in {FW} and HWVersion in {HW} and BoardNo in {Board} and DUTName in {DUTname} and DUTID in {DUTID} and
                # ChapterName in {Chap} and Coil in {Coil} and TestcaseID in {Tests}
                # ORDER by Header.TestcaseName,ChecksHeader.SEQID,ChecksHeader.SEQID, ChecksDetails.CheckSEQ
                # ChecksHeader.Type="Measures" and
                Header_df = pd.read_sql_query(Header_Qry, self.connection)
                # print(Header_df)
                if product=="MPP":
                    Header_df["Flow"] = Header_df["Flow"].replace("1", "128kHz")
                    Header_df["Flow"] = Header_df["Flow"].replace("2", "360kHz")
                else: Header_df["Flow"] = Header_df["Flow"]
                # print(Header_df)
                # Create a pivot table
                pivot_table = pd.pivot_table(
                    Header_df,
                    index=['TestcaseName','Flow','CheckType','PacketID','Packet','HeaderName','HeaderResult','Byte','Bit','CheckName','ExpValue','RecValue','ChecksResult'],  # Rows
                    values=[],        # Values to aggregate
                    aggfunc='sum'                        # Aggregation function
                )
                FinalDataFrames[pha]=pivot_table
                # print(pivot_table)
            
            now = datetime.now()
            timestamp = now.strftime("%d%m%Y_%H%M%S")
            output_file = f'Results/C3_MPP Excel Results/PayLoadChecks_Report_{filters['Product']}_{filters['Certification']}_{filters['SW'][0]}_{filters['FW'][0]}_{timestamp}.xlsx'
            with pd.ExcelWriter(output_file) as writer:
                for Phases in FinalDataFrames:
                    phasename = Phases if len(Phases)<30 else Phases[0:30]
                    FinalDataFrames[Phases].to_excel(writer, sheet_name=phasename)
                # for sheet_name, df in FinalDataFrames.items():
                    # df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            #     pivot_table.to_excel(writer, sheet_name='Pivot Table')
            #Update Excel sheet
            self.format_excel(output_file)
    
    def format_excel(self,file_path):
        # Define color fills for results (Global to avoid redundant creation)
        COLOR_MAPPING = {
            "pass": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Green
            "fail": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
            "inconclusive": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Yellow
            "na": PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid"),
        }
        
        """Loads, processes, and saves the Excel file with formatting."""
        # Load the Excel file
        wb = load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]  # Get the active sheet
            # print("sheet_name:",sheet_name)

            # Auto-fit all columns based on content
            for col in ws.iter_cols():
                # print("col:",col)
                col_letter = col[0].column_letter  # Get column letter
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[col_letter].width = max_length + 2  # Add padding

            # Find the "Remarks" column dynamically
            remarks_col_index = next(
                (col[0].column for col in ws.iter_cols(1, ws.max_column) 
                if col[0].value and str(col[0].value).strip().lower() == "remarks"),
                None
            )

            # Align "Remarks" column to left if found
            if remarks_col_index:
                for cell in ws.iter_cols(min_col=remarks_col_index, max_col=remarks_col_index, 
                                        min_row=2, max_row=ws.max_row):
                    for c in cell:
                        c.alignment = Alignment(horizontal="left")

            # Apply color coding for Pass/Fail/Inconclusive
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value).strip().lower()
                        if cell_value in COLOR_MAPPING:
                            cell.fill = COLOR_MAPPING[cell_value]

        # Save the updated file
        wb.save(file_path)

# obj = ExcelReports()
# obj.CTSChecksReport()

# obj = ExcelReports()
# obj.CTSDetailedReport()
# flr = {'SW': ['2.220.0.39'], 'FW': ['4.0.3.35'], 'HW': ['E-3.9'], 'Board': ['GRL-C3-MP-2024113'], 'DUTname': [''], 'DUTID': [''], 'Chap': ['Digital_Ping', 'K_est_requirements'], 'Coil': ['NA'], 'Tests': ['MPP_PTX_POW_Digital_Ping_360_LPM_TC3', 'MPP_PTX_POW_Digital_Ping_360_LPM_TC4', 'MPP_PTX_POW_Digital_Ping_360_NPM_TC1', 'MPP_PTX_POW_Digital_Ping_360_NPM_TC2', 'MPP_PTX_POW_Digital_Ping_360_NPM_TC3', 'MPP_PTX_POW_Digital_Ping_360_NPM_TC4', 'MPP_PTX_POW_Digital_Ping_360_HPM_TC1', 'MPP_PTX_POW_Digital_Ping_360_LPM_TC2', 'MPP_PTX_POW_Digital_Ping_360_HPM_TC2', 'MPP_PTX_POW_Digital_Ping_360_HPM_TC4', 'MPP_PTX_POW_Digital_Ping_360_CPM_TC1', 'MPP_PTX_POW_Digital_Ping_360_CPM_TC2', 'MPP_PTX_POW_Digital_Ping_360_CPM_TC3', 'MPP_PTX_POW_Digital_Ping_360_CPM_TC4', 'MPP_PTX_POW_Digital_Ping_360_OV_LPM_TC1', 'MPP_PTX_POW_Digital_Ping_360_OV_LPM_TC2', 'MPP_PTX_POW_Digital_Ping_360_HPM_TC3', 'MPP_PTX_POW_Digital_Ping_360_LPM_TC1', 'MPP_PTX_POW_Digital_Ping_128_HPM_Reping', 'MPP_PTX_POW_Digital_Ping_128_HPM_P2', 'MPP_PTX_POW_Cloak_Ping_360_LPM_TC1', 'MPP_PTX_POW_Cloak_Ping_360_LPM_TC2', 'MPP_PTX_POW_Cloak_Ping_360_LPM_TC3', 'MPP_PTX_POW_Cloak_Ping_360_LPM_TC4', 'MPP_PTX_POW_Cloak_Ping_360_NPM_TC1', 'MPP_PTX_POW_Cloak_Ping_360_NPM_TC2', 'MPP_PTX_POW_Cloak_Ping_360_NPM_TC3', 'MPP_PTX_POW_Cloak_Ping_360_NPM_TC4', 'MPP_PTX_POW_Cloak_Ping_360_HPM_TC1', 'MPP_PTX_POW_Cloak_Ping_360_HPM_TC2', 'MPP_PTX_POW_Cloak_Ping_360_HPM_TC3', 'MPP_PTX_POW_Cloak_Ping_360_HPM_TC4', 'MPP_PTX_POW_Cloak_Ping_360_CPM_TC1', 'MPP_PTX_POW_Cloak_Ping_360_CPM_TC2', 'MPP_PTX_POW_Cloak_Ping_360_CPM_TC3', 'MPP_PTX_POW_Cloak_Ping_360_CPM_TC4', 'MPP_PTX_POW_Digital_Ping_128_HPM_P1', 'MPP_PTX_POW_Digital_Ping_360_OV_NPM_TC1', 'MPP_PTX_POW_Digital_Ping_360_OV_NPM_TC2', 'MPP_PTX_POW_Digital_Ping_360_OV_HPM_TC1', 'MPP_PTX_POW_Digital_Ping_360_OV_HPM_TC2', 'MPP_PTX_POW_Digital_Ping_360_OV_CPM_TC1', 'MPP_PTX_POW_Digital_Ping_360_OV_CPM_TC2', 'MPP_PTX_POW_Cloak_Ping_360_OV_LPM_TC1', 'MPP_PTX_POW_Cloak_Ping_360_OV_LPM_TC2', 'MPP_PTX_POW_Cloak_Ping_360_OV_NPM_TC1', 'MPP_PTX_POW_Cloak_Ping_360_OV_NPM_TC2', 'MPP_PTX_POW_Cloak_Ping_360_OV_HPM_TC1', 'MPP_PTX_POW_Cloak_Ping_360_OV_HPM_TC2', 'MPP_PTX_POW_Cloak_Ping_360_OV_CPM_TC1', 'MPP_PTX_POW_Cloak_Ping_360_OV_CPM_TC2', 'MPP_PTX_POW_KEst_P1', 'MPP_PTX_POW_KEst_P2', 'MPP_PTX_POW_KEst_HPM_P1', 'MPP_PTX_POW_KEst_HPM_P2', 'MPP_PTX_NEG_POW_KEST_SLIDING'], 'Timings': False, 'Measures': False, 'Others': False, 'Product': 'MPP_TPR', 'Certification': '2.2.0'}
# obj.CTSDetailedReport(flr)
# obj.CTSChecksReport(flr)V22_Final_TestBackup.gproj

# pivot_table = merged_df.pivot_table(values='Description', index='Product', columns='SEQID', fill_value=0)
# print(pivot_table)
# Display the DataFrame
# print(Measures_df)



# import pandas as pd

# # Sample data for the pivot table
# data = {
#     'Category': ['A', 'B', 'A', 'C', 'B', 'A', 'C', 'C'],
#     'Sub-Category': ['X', 'Y', 'X', 'Z', 'Y', 'Z', 'X', 'Y'],
#     'Sales': [100, 200, 150, 300, 400, 500, 250, 450],
#     'Quantity': [1, 2, 3, 4, 2, 5, 3, 2]
# }

# # Create DataFrame
# df = pd.DataFrame(data)

# # Create a pivot table
# pivot_table = pd.pivot_table(
#     df,
#     index=['Category', 'Sub-Category'],  # Rows
#     values=['Sales', 'Quantity'],        # Values to aggregate
#     aggfunc='sum'                        # Aggregation function
# )

# # Define output Excel file
# output_file = 'PivotTable_Report.xlsx'

# # Write both the original data and the pivot table to an Excel file
# with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
#     # Write the original data
#     df.to_excel(writer, sheet_name='Data', index=False)
    
#     # Write the pivot table
#     pivot_table.to_excel(writer, sheet_name='Pivot Table')

# print(f"Pivot table report created and saved as '{output_file}'.")
