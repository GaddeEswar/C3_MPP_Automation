import socket
import threading
import time,requests
import subprocess
from scapy.all import sr1, IP, ICMP
import psutil
from itertools import combinations
from datetime import datetime, timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from MainModule import JsonOperations,APIOperations
from tkinter import messagebox

class GrlEthernetLink_C2:
    def __init__(self, PermLicense, Perm_DemoLicense, DemoLicense, mode):
        self.Mode = mode
        self.is_port_open = False
        self.use_ethernet_link = True
        self.m_databuffersize = 1024
        self.m_read_port = 5002
        self.m_write_port = 5002
        self.m_debug_write = True
        self.m_debug_read = False
        self.i_write_time_delay = 10
        self.i_read_time_delay = 15
        self.m_ip_address = "192.168.255.1"
        self.m_read_client = None
        self.m_write_client = None
        self.m_read_stream = None
        self.m_write_stream = None
        self.m_read_retry_count = 0
        self.m_write_retry_count = 0
        self.m_is_app_forced_to_stop = True
        self.m_eth_ip = None
        self.m_eth_dns = None
        self.m_eth_name = None
        self.onebyte = 1
        self.twobytes = 2
        self.JLogs = JsonOperations("json/DebugLogs.json")
        self.JLogsData = self.JLogs.read_file()
        self.License = JsonOperations("json/LicenseCheck.json")
        self.LicenseData = self.License.read_file()
        self.Jsettings = JsonOperations('json/setting.json')
        self.Jtester = JsonOperations('json/Tester.json')
        self.JtesterData =self.Jtester.read_file()

        # self.UpdatePermLicense = PermLicense
        # self.UpdateTempLicense = DemoLicense
        self.PermLicense = PermLicense
        self.Perm_DemoLicense = Perm_DemoLicense
        self.DemoLicense = DemoLicense
        self.TesterIP = "192.168.255.1"
        self.SerialNumber = None
        
        self.m_read_write_lock = threading.Lock()
        self.m_read_api_lock = threading.Lock()

    def ethernet_discovery(self):
        # Discover network interfaces and select the Ethernet interface
        interfaces = psutil.net_if_addrs()
        ip = dns = nic = None
        for interface, addrs in interfaces.items():
            
            for addr in addrs:
                #print(addr)
                #if addr.family == psutil.AF_LINK:  # Check for IPv4
                if addr.family == socket.AF_INET:
                    ip_addr = addr.address
                    if not ip_addr.startswith("169"):  # Exclude automatic IPs
                        ip = ip_addr
                        nic = interface
                        break
            if ip:  # Break out of the outer loop if IP is found
                break
        
        return ip, dns, nic
    
    def set_ip(self, ip):
        # Set IP using netsh (Windows-specific, requires admin)
        command = f'netsh interface ip set address "{self.m_eth_name}" static {ip}'
        #print(f"Running command: {command}")  # Debugging line
        try:
            subprocess.run(command, shell=True, check=True)
            print(f"Running command: {command}")  # Debugging line
        except subprocess.CalledProcessError:
            print("Failed to set IP")

    def initialize_port(self):
        try:
            if not self.is_port_open:
                self.m_eth_ip, self.m_eth_dns, self.m_eth_name = self.ethernet_discovery()
                #print(self.m_eth_ip, self.m_eth_dns, self.m_eth_name)
                if self.m_eth_ip == self.m_ip_address:
                    print("hi")
                    # ip_parts = self.m_eth_ip.split('.')
                    # last_part = int(ip_parts[-1]) + 1
                    # ip_parts[-1] = str(last_part)
                    # self.m_eth_ip = '.'.join(ip_parts)
                    self.set_ip(self.m_eth_ip)
                    print("setting:",self.m_eth_ip)

                self.m_read_client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                self.m_read_client.connect((self.m_ip_address, self.m_read_port))
                self.m_read_client.settimeout(20)

                #print(self.m_read_client)

                self.m_write_client = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                self.m_write_client.connect((self.m_ip_address, self.m_write_port))
                self.m_write_client.settimeout(20)
                
                #print(self.m_read_client,self.m_write_client)
                if self.m_read_client and self.m_write_client:
                    #print(self.m_read_client,self.m_write_client)
                    self.is_port_open = True
                    self.m_is_app_forced_to_stop = False
        except Exception as ex:
            print(f"Error initializing port: {ex}")
            self.is_port_open = False
            return False
        return True

    def close_link(self):
        self.dispose()

    def write(self, buffer):
        #self.size = size
        with self.m_read_api_lock:
            if self.m_write_client:
                try:
                    if self.m_write_client:
                        self.m_write_client.sendall(buffer)
                        #   time.sleep(self.i_write_time_delay / 1000)
                        # response = self.m_write_client.recv(20)    #response
                        print("License written")
                        # #print('Received:', list(response))#
                        return True
                    else:
                        self.retry_write_operation(buffer)
                except Exception as ex:
                    print(f"Write error: {ex}")
                    self.retry_write_operation(buffer)
        return False

    def retry_write_operation(self, buffer):
        if self.m_write_retry_count > 10:
            return False
        self.m_write_retry_count += 1
        time.sleep(1)
        self.initialize_port()
        return self.write(buffer)

    def dispose(self):
        try:
            if self.m_read_client:
                self.m_read_client.close()
            if self.m_write_client:
                self.m_write_client.close()
            self.is_port_open = False
        except Exception as ex:
            print(f"Error disposing connection: {ex}")
            self.is_port_open = False

    def MPPLicenseUpdate(self,address,License,Noofbytes):
        self.initialize_port()
        self.bufferdata = bytearray([0x03, 0x02, 0x01, 0xA0, 0x00, 0x00, 0x00, 0x00])
        self.bufferdata[5] = Noofbytes + 2
        self.bufferdata[6] = (address >> 8) & 0xFF
        self.bufferdata[7] = address & 0xFF

        self.bufferdata.extend(License.to_bytes((License.bit_length()+7)//8,'little'))
        print(bytes(self.bufferdata))

        #self.write(b'\xcd\x03\x00\x00\x04\x00')  # Write to read device data
        self.write(bytes(self.bufferdata))   #Writing the License
        self.is_port_open = False

    def PreExecute(self):
        #print(f"Perm:{self.UpdatePermLicense}, TEMP:{self.UpdateTempLicense}, MODE:{self.Mode}")
        self.isMPPBoard = True 

        #Perm License FRAM Address
        self.PermLicenseAdd = 0x0050
        #Perm usage License FRAM Address          
        self.PermUsageLicenseAdd = 0x0052 

        # Temp License start data FRAM addresses
        self.StartDateAdd = 0x0444
        self.StartMonthAdd = 0x0445
        self.StartYearAdd = 0x0446
        # Temp License stop data FRAM addresses
        self.StopDateAdd = 0x0448
        self.StopMonthAdd = 0x0449
        self.StopYearAdd = 0x044A
        # Temp License FRAM Address
        self.TempLicenseAdd = 0x0442  

        presentday = datetime.now().strftime('%d-%m-%Y').split("-")
        # Temp License start data
        self.SetStartDate = int(presentday[0])#29#0x1D
        self.SetStartMonth = int(presentday[1])#10#0x0A
        self.SetStartYear = int(presentday[2])#2024#0x7E8
        tomorrow = (datetime.now()+timedelta(1)).strftime('%d-%m-%Y').split("-")
        # Temp License stop data
        self.SetEndDate = int(tomorrow[0])#30#0x1E
        self.SetEndMonth = int(tomorrow[1])#10#0x0A
        self.SetEndYear = int(tomorrow[2])

        self.LicenseComb()

    # Get all License Combinations
    def LicenseComb(self):
        #[1,100,200,400,800,1000]  MPP-TPR License values, # [100,200,400,800,1,2000,1000,4000,10000,20000,40000,80000,100000] MPP-TPT License values
        data = [1, 256, 512, 1024, 2048, 4096] if self.Mode == "TPR" else [256, 512, 1024, 2048, 1, 8192, 4096, 16384, 65536, 131072, 262144, 524288, 1048576]
        #data = [1, 256] if self.Mode == "TPR" else [256, 512, 1024, 2048, 1, 8192, 4096, 16384, 65536, 131072, 262144, 524288, 1048576]
        n = len(data)
        Liccomblst = []

        if self.Mode == "TPR":
            for r in range(1, n + 1):
                for combination in combinations(data, r):
                    remaining_elements = [item for item in data if item not in combination]

                    # Calculate sums
                    comb_sum = sum(combination)
                    remaining_sum = sum(remaining_elements)

                    # Convert data elements and sums to hex with 4 digits, zero-padded
                    combination_hex = [f"0x{item:04x}" for item in combination]
                    remaining_hex = [f"0x{item:04x}" for item in remaining_elements]
                    comb_sum_hex = f"0x{comb_sum:04x}"
                    remaining_sum_hex = f"0x{remaining_sum:04x}"

                    # print(f"Comb: {combination_hex}, Sum: {comb_sum_hex}, "
                    #     f"Remaining: {remaining_hex}, Sum of remaining: {remaining_sum_hex}")

                    Liccomblst.append([comb_sum_hex, remaining_sum_hex])
        else:
            if self.Perm_DemoLicense:
                Liccomblst = [[0x00007F01, 0x001F0000]]
            elif self.PermLicense:
                Liccomblst = [[0x001F7F01, 0X00000000], [0x00107F01, 0x000F0000], [0x00000300, 0x001F7C01], [0x00000F00, 0x001F7001], [0x00007F01, 0x001F0000], [0x001F7000, 0x00000F01]]
        self.LicenseVerify(Liccomblst)

    def replace_obtained_result(self, obj):
        if isinstance(obj, dict):
            for key in list(obj.keys()):
                if key == "ObtainedResult":
                    obj[key] = ""
                else:
                    self.replace_obtained_result(obj[key])
        elif isinstance(obj, list):
            for item in obj:
                self.replace_obtained_result(item)

    def LicenseVerify(self, Liccomblst):
        try:
            # Writing temp license start and stop data
            if self.Perm_DemoLicense or self.DemoLicense:
                self.MPPLicenseUpdate(self.StartDateAdd,self.SetStartDate,self.onebyte)
                self.MPPLicenseUpdate(self.StartMonthAdd,self.SetStartMonth,self.onebyte)
                
                self.MPPLicenseUpdate(self.StartYearAdd,self.SetStartYear,self.twobytes)
                self.MPPLicenseUpdate(self.StopDateAdd,self.SetEndDate,self.onebyte)

                self.MPPLicenseUpdate(self.StopMonthAdd,self.SetEndMonth,self.onebyte)
                self.MPPLicenseUpdate(self.StopYearAdd,self.SetEndYear,self.twobytes)

            # Delete previous results
            self.replace_obtained_result(self.LicenseData)

            self.LicType = {"PERM":self.PermLicense, "PERM_DEMO":self.Perm_DemoLicense, "DEMO":self.DemoLicense}
            for litype in self.LicType:
                self.Lic = None
                i = 1
                if self.LicType[litype]:
                    self.LicenseData["MPP"][self.Mode][litype]["SerialNumber"] = self.JtesterData["MPP"][self.Mode]["BoardNo"]
                    self.LicenseData["MPP"][self.Mode][litype]["SoftwareVersion"] = self.JtesterData["MPP"][self.Mode]["SWVersion"]
                    self.LicenseData["MPP"][self.Mode][litype]["Mode"] = self.Mode
                    for lic in Liccomblst:
                        #Force stopping
                        self.JsettingsData =  self.Jsettings.read_file()
                        if self.JsettingsData['_stop_flag']:
                            self.update_logs("UI", "Force stopping License Verification")
                            raise LoopExit
                        
                        self.update_logs("UI", f"License Verifying : {litype} - Comb {i}/{len(Liccomblst)} - {lic}")
                        if litype == "PERM" and self.LicType[litype]:
                            self.SetPermanetLicense = int(lic[0],16)
                            #self.data[f"Comb {i}"]["License"] = f"PERM: {lic[0]}"
                            self.Lic = {"PERM": f"PERM: {lic[0]}"}

                        elif litype == "PERM_DEMO" and self.LicType[litype]:
                            self.SetPermanetLicense = int(lic[0],16) | 32768
                            self.SetTemporaryLicense = int(lic[1],16)
                            #self.data[f"Comb {i}"]["License"] = f"PERM: {lic[0]},DEMO: {lic[1]}"
                            self.Lic = {"PERM_DEMO": f"PERM: {lic[0]},DEMO: {lic[1]}"}

                        elif litype == "DEMO" and self.LicType[litype]:
                            self.SetPermanetLicense = 0 | 32768
                            self.SetTemporaryLicense = int(lic[0],16)
                            #self.data[f"Comb {i}"]["License"] = f"DEMO: {lic[0]}"
                            self.Lic = {"DEMO": f"DEMO: {lic[0]}"}

                        # Write Permanent License
                        self.MPPLicenseUpdate(self.PermLicenseAdd,self.SetPermanetLicense,self.twobytes) 
                        self.SetPermanentUsageLicense = self.SetPermanetLicense  
                        self.MPPLicenseUpdate(self.PermUsageLicenseAdd,self.SetPermanentUsageLicense,self.twobytes)
                        
                        # Write Temporary License
                        if litype in ["PERM_DEMO","DEMO"]: self.MPPLicenseUpdate(self.TempLicenseAdd,self.SetTemporaryLicense,self.twobytes)
                        time.sleep(2)

                        # Powercycle the tester
                        self.Reboot()
                        time.sleep(80) # Average tester booting time

                        # Ping the tester
                        count_ping = 0
                        while count_ping < 60:
                            print("UI",f"Pinging {self.TesterIP}...")
                            count_ping += 1
                            time.sleep(1)
                            if self.ping_ip() == True:
                                self.update_logs("Connection",f"{self.TesterIP} IP is reached")
                                print("UI",f"{self.TesterIP} IP is reached")
                                time.sleep(5)
                                tester_data = self.TesterConnect() 
                                if self.LicenseData["MPP"][self.Mode][next(iter(self.Lic))][f"Comb {i}"]["License"] == self.Lic[next(iter(self.Lic))]:
                                    self.LicenseData["MPP"][self.Mode][next(iter(self.Lic))][f"Comb {i}"]["ObtainedResult"] = tester_data['licenseInfo']
                                self.SerialNumber = tester_data['serialNumber']
                                break
                            if count_ping >= 60:
                                self.update_logs("UI",f"Please check Ethernet connection")
                                print("UI",f"Please check Ethernet connection")
                                raise LoopExit
                        print(f"License update for {int(str(lic[0]),16),int(str(lic[1]),16)} ended")
                        self.is_port_open = False
                        i += 1
                    self.License.update_file(self.LicenseData)
        except LoopExit:
            self.License.update_file(self.LicenseData)
            # self.EnableAllLicenses()
            pass
        self.convert_data_to_dataframe()
        # self.EnableAllLicenses()

    #Enable all permanent licenses after execution
    def EnableAllLicenses(self):
        # Write Permanent License
        self.MPPLicenseUpdate(self.PermLicenseAdd,self.SetPermanetLicense,self.twobytes) 
        self.SetPermanentUsageLicense = self.SetPermanetLicense  
        self.MPPLicenseUpdate(self.PermUsageLicenseAdd,self.SetPermanentUsageLicense,self.twobytes)

        time.sleep(2)
        # Powercycle the tester
        self.Reboot()

    #Tester reboot
    def Reboot(self):
        command = bytes.fromhex("03 02 01 30 00 02 AA 03")
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.connect((self.m_ip_address, self.m_write_port))
                s.sendall(command)
                print(" TCP reboot command sent.")
                # response = s.recv(1024)
                # print(" Received response:", response.hex())
            except Exception as e:
                print(" Error sending TCP command:", e)

    def convert_data_to_dataframe(self):
        self.update_logs("UI",f"Report generating")
        wb = Workbook()
        # Remove the default sheet created by openpyxl
        default_sheet = wb.active
        wb.remove(default_sheet)

        for lic in self.LicenseData["MPP"][self.Mode]:
            data = self.LicenseData["MPP"][self.Mode][lic]

            # Check if any combination has a non-empty ObtainedResult
            has_valid_comb = any(
                data[key].get("ObtainedResult")
                for key in data if key.startswith("Comb")
            )

            if not has_valid_comb:
                continue  # Skip sheet creation if no valid combinations

            ws = wb.create_sheet(title=lic[:31])  # Sheet name max length = 31

            # Add metadata
            ws['A1'] = 'SerialNumber'
            ws['B1'] = data['SerialNumber']
            ws['A2'] = 'SoftwareVersion'
            ws['B2'] = data['SoftwareVersion']
            ws['A3'] = 'Mode'
            ws['B3'] = data['Mode']

            # Styles
            header_fill = PatternFill(start_color="F0C870", end_color="F0C870", fill_type="solid")
            pass_fill = PatternFill(start_color="21EA49", end_color="21EA49", fill_type="solid")
            fail_fill = PatternFill(start_color="F80B0B", end_color="F80B0B", fill_type="solid")

            # Headers
            header_row = 5
            headers = ['S.No', 'Combinations', 'License Type', 'Expected Result', 'Obtained Result', 'Result']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = header_fill

            # Fill data
            row = header_row + 1
            combo_number = 1
            for key in sorted([k for k in data if k.startswith('Comb')], key=lambda x: int(x.split()[-1])):
                comb = data[key]
                ExpectedResult = comb.get('ExpectedResult', [])
                ObtainedResult = comb.get('ObtainedResult', [])

                if not ObtainedResult:
                    continue  # Skip this combination if ObtainedResult is empty

                match_count = len(ExpectedResult)

                all_match = all(
                    exp['moduleName'] == obt['moduleName'] and exp['moduleStatus'] == obt['moduleStatus']
                    for exp, obt in zip(ExpectedResult, ObtainedResult)
                )

                for i in range(match_count):
                    ws.cell(row=row + i, column=3, value=ExpectedResult[i]['moduleName'])
                    ws.cell(row=row + i, column=4, value=ExpectedResult[i]['moduleStatus'])
                    ws.cell(row=row + i, column=5, value=ObtainedResult[i]['moduleStatus'])

                ws.cell(row=row, column=1, value=combo_number)
                ws.cell(row=row, column=2, value=comb['License'])
                ws.merge_cells(start_row=row, start_column=1, end_row=row + match_count - 1, end_column=1)
                ws.merge_cells(start_row=row, start_column=2, end_row=row + match_count - 1, end_column=2)
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='top')
                ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')

                result = "PASS" if all_match else "FAIL"
                result_cell = ws.cell(row=row, column=6, value=result)
                ws.merge_cells(start_row=row, start_column=6, end_row=row + match_count - 1, end_column=6)
                result_cell.fill = pass_fill if result == "PASS" else fail_fill
                result_cell.alignment = Alignment(horizontal='center', vertical='center')

                row += match_count + 1
                combo_number += 1

            # Set column widths
            col_widths = [20, 35, 25, 20, 20, 10]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Freeze header
            ws.freeze_panes = 'A6'

        # Protect the worksheet from editing
        ws.protection.sheet = True
        ws.protection.password = '1234'  # Set a password

        # Save workbook
        now = datetime.now()
        timestamp = now.strftime("%d%m%Y_%H%M%S")
        wb.save("Results/License Results/License_verification_"+ f"{self.SerialNumber}_"+ timestamp + '.xlsx')
        
    def Connpopup(self):
         while not self.Conn_flag:
            popupdata = {"userTextBoxInput":"","responseButton":"Ok","shouldTextBoxBeAdded":False,"isValid":True,"popID":15,"displayPopUp":False,"isDisplayPopUpOpen":False,"title":"GRL-C3-MP-TPR Test Solution","message":"The controller calibration was going to expire in 19 days. Please contact GRL support.!","button":"OK","image":None,"icon":"Asterisk","isFrontEndPopUp":False,"callBackMethod":None,"comboBoxEntries":None,"selectedComboBoxValue":"","comboBoxEntriesFE":[],"selectedComboBoxValueFE":"","onlyDropdownAdded":False,"enableTimerOKButton":False,"enableCustomUserInputs":False,"customInputValues":{}}
            requests.put("http://localhost:2002/api/App/PutMessageBoxResponse", json = popupdata)

    def TesterConnect(self):
        self.Conn_flag = False
        time.sleep(1)
        threading.Thread(target=self.Connpopup,daemon=True).start()
        requests.put("http://localhost:2002/api/App/PutApplicationActiveStatus/true")
        testerinfo = requests.get("http://localhost:2002/api/ConnectionSetup/192.168.255.1").json()
        if testerinfo is not None:
            if testerinfo['testerStatus'] == 'Connected':
                self.Conn_flag = True
                print(testerinfo)
                return testerinfo
        else: 
            self.Conn_flag = True
            time.sleep(2)
            self.TesterConnect()
        
    def ping_ip(self):
        packet = IP(dst=self.TesterIP)/ICMP()
        start = time.time()
        reply = sr1(packet, timeout=1, verbose=0)
        if reply:
            return True
        else:
            return False

    def update_logs(self,logtype,log):
        dt_object = datetime.fromtimestamp(datetime.now().timestamp())
        self.JLogsData = self.JLogs.read_file()
        self.JLogsData.append([str(dt_object),logtype,log])
        self.JLogs.update_file(self.JLogsData)
        
class LoopExit(Exception):
    pass

if __name__ == "__main__":
   
    ethernet_link = GrlEthernetLink_C2(True,True,True, 'TPR')
#     # ethernet_link.PreExecute()
#     #ethernet_link.convert_data_to_dataframe({'Comb 1': {'License': 'PERM: 0x0001, DEMO: 0x0100', 'Testerinfo': [{'moduleName': 'MPP 2.0', 'moduleStatus': 'ACTIVE'}, {'moduleName': 'MPP 2.1', 'moduleStatus': 'DEMO'}, {'moduleName': 'APP 2.1', 'moduleStatus': 'NOT_AVAILABLE'}, {'moduleName': '2.2 - MPP 25W', 'moduleStatus': 'NOT_AVAILABLE'}, {'moduleName': '2.2 - MPP 15W', 'moduleStatus': 'NOT_AVAILABLE'}, {'moduleName': '2.2 - APP 15W & 25W', 'moduleStatus': 'NOT_AVAILABLE'}]}, 'SerialNumber': 'GRL-C3-MP-2019023', 'SoftwareVersion': '2.220.0.47 - 25W-Beta', 'Mode': 'TPR','Comb 2': {'License': 'PERM: 0x0100, DEMO: 0x0001', 'Testerinfo': [{'moduleName': 'MPP 2.0', 'moduleStatus': 'DEMO'}, {'moduleName': 'MPP 2.1', 'moduleStatus': 'ACTIVE'}, {'moduleName': 'APP 2.1', 'moduleStatus': 'NOT_AVAILABLE'}, {'moduleName': '2.2 - MPP 25W', 'moduleStatus': 'NOT_AVAILABLE'}, {'moduleName': '2.2 - MPP 15W', 'moduleStatus': 'NOT_AVAILABLE'}, {'moduleName': '2.2 - APP 15W & 25W', 'moduleStatus': 'NOT_AVAILABLE'}]}, 'Comb 3': {'License': 'PERM: 0x0101, DEMO: 0x0000', 'Testerinfo': [{'moduleName': 'MPP 2.0', 'moduleStatus': 'ACTIVE'}, {'moduleName': 'MPP 2.1', 'moduleStatus': 'ACTIVE'}, {'moduleName': 'APP 2.1', 'moduleStatus': 'NOT_AVAILABLE'}, {'moduleName': '2.2 - MPP 25W', 'moduleStatus': 'NOT_AVAILABLE'}, {'moduleName': '2.2 - MPP 15W', 'moduleStatus': 'NOT_AVAILABLE'}, {'moduleName': '2.2 - APP 15W & 25W', 'moduleStatus': 'NOT_AVAILABLE'}]}})
    ethernet_link.convert_data_to_dataframe()